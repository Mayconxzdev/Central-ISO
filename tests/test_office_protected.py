from pathlib import Path
import tempfile

import pytest
from openpyxl import Workbook

from app.services.extractors import extract_text


pytest.importorskip("msoffcrypto")


def make_encrypted_xlsx(path: Path, password: str) -> None:
    from msoffcrypto.format.ooxml import OOXMLFile

    plain = path.with_name("plain.xlsx")
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Codigo"
    ws["B1"] = "Valor protegido de teste"
    wb.save(plain)
    with open(plain, "rb") as source, open(path, "wb") as output:
        OOXMLFile(source).encrypt(password, output)
    plain.unlink()


def test_encrypted_office_file_is_read_with_configured_password(tmp_path: Path, monkeypatch):
    encrypted = tmp_path / "protected.xlsx"
    make_encrypted_xlsx(encrypted, "unit-test-password")
    monkeypatch.setenv("ISO_PROTECTED_FILE_PASSWORD", "unit-test-password")

    text, status = extract_text(encrypted)

    assert status == "protegido e lido"
    assert "Valor protegido de teste" in text


def test_encrypted_office_file_rejects_wrong_password_and_removes_temp(tmp_path: Path, monkeypatch):
    encrypted = tmp_path / "protected.xlsx"
    make_encrypted_xlsx(encrypted, "unit-test-password")
    monkeypatch.setenv("ISO_PROTECTED_FILE_PASSWORD", "wrong-password")

    text, status = extract_text(encrypted)

    assert text == ""
    assert status == "protegido, senha nao funcionou"
    temp_root = Path(tempfile.gettempdir())
    assert not list(temp_root.glob("central_iso_office_*"))
